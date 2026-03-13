import os
import csv
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from scr.py_to_html import _save_html


def save_to_files(posts):
    """Save list of posts to XLSX, CSV, TXT, and HTML files"""
    if not posts:
        return

    today_date = datetime.date.today().strftime("%Y-%m-%d")
    folder = f"data/{today_date}"
    os.makedirs(folder, exist_ok=True)

    _save_xlsx(posts, folder)
    _save_csv(posts, folder)
    _save_txt(posts, folder)
    _save_html(posts, folder, today_date)

    print(f" All files saved to {folder}/")


# ─────────────────────────────────────────────
# XLSX
# ─────────────────────────────────────────────
def _save_xlsx(posts, folder):
    wb = Workbook()
    ws = wb.active
    ws.title = "Startup Ideas"

    header_font     = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill     = PatternFill("solid", fgColor="1F4E79")
    header_align    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font      = Font(name="Arial", bold=True, color="1F4E79", size=14)
    subtitle_font   = Font(name="Arial", italic=True, color="666666", size=10)
    row_fill_even   = PatternFill("solid", fgColor="EBF3FB")
    row_fill_odd    = PatternFill("solid", fgColor="FFFFFF")
    row_font        = Font(name="Arial", size=10)
    row_align_wrap  = Alignment(vertical="top", wrap_text=True)
    row_align_center= Alignment(horizontal="center", vertical="top")
    thin_border     = Border(
        left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),  bottom=Side(style="thin", color="D9D9D9"),
    )
    upvote_font_high = Font(name="Arial", bold=True, color="1E7E34", size=10)
    upvote_font_mid  = Font(name="Arial", bold=True, color="856404", size=10)
    upvote_font_low  = Font(name="Arial", size=10, color="555555")

    ws.merge_cells("A1:C1")
    ws["A1"] = " Startup Ideas Dashboard"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Generated on {datetime.datetime.now().strftime('%B %d, %Y  %H:%M')}  •  {len(posts)} ideas"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    for col, header in enumerate(["Startup Name", "Description", "Upvotes"], start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[3].height = 28

    for i, (title, description, upvotes) in enumerate(posts, start=1):
        row = i + 3
        fill = row_fill_even if i % 2 == 0 else row_fill_odd

        c_name = ws.cell(row=row, column=1, value=title)
        c_name.font = Font(name="Arial", bold=True, size=10)
        c_name.fill = fill; c_name.alignment = row_align_wrap; c_name.border = thin_border

        c_desc = ws.cell(row=row, column=2, value=description)
        c_desc.font = row_font
        c_desc.fill = fill; c_desc.alignment = row_align_wrap; c_desc.border = thin_border

        c_up = ws.cell(row=row, column=3, value=upvotes)
        c_up.fill = fill; c_up.alignment = row_align_center; c_up.border = thin_border
        if isinstance(upvotes, int):
            c_up.font = upvote_font_high if upvotes >= 500 else (upvote_font_mid if upvotes >= 100 else upvote_font_low)
        else:
            c_up.font = row_font
        ws.row_dimensions[row].height = 40

    summary_row = len(posts) + 4
    for col, val in [(1, "TOTAL IDEAS"), (2, f"=COUNTA(A4:A{summary_row-1})"), (3, f"=SUM(C4:C{summary_row-1})")]:
        cell = ws.cell(row=summary_row, column=col, value=val)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[summary_row].height = 24

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:C{summary_row - 1}"

    wb.save(f"{folder}/startup_ideas.xlsx")


# ─────────────────────────────────────────────
# CSV
# ─────────────────────────────────────────────
def _save_csv(posts, folder):
    with open(f"{folder}/startup_ideas.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        writer.writerow(["#", "Startup Name", "Description", "Upvotes", "Generated Date"])
        generated = datetime.date.today().strftime("%Y-%m-%d")
        for i, (title, description, upvotes) in enumerate(posts, start=1):
            writer.writerow([i, title, description, upvotes, generated])
        # Summary footer
        writer.writerow([])
        writer.writerow(["", "TOTAL IDEAS", len(posts), sum(u for _, _, u in posts if isinstance(u, int)), ""])


# ─────────────────────────────────────────────
# TXT
# ─────────────────────────────────────────────
def _save_txt(posts, folder):
    col_w = [4, 38, 52, 8]
    sep   = "+" + "+".join("-" * (w + 2) for w in col_w) + "+"

    def row_line(values, widths, bold=False):
        cells = []
        for v, w in zip(values, widths):
            v = str(v)
            v = (v[:w-3] + "...") if len(v) > w else v
            cells.append(f" {v:<{w}} ")
        return "|" + "|".join(cells) + "|"

    with open(f"{folder}/startup_ideas.txt", "w", encoding="utf-8") as f:
        now = datetime.datetime.now().strftime("%B %d, %Y %H:%M")
        f.write("=" * (sum(col_w) + len(col_w) * 3 + 1) + "\n")
        f.write(f"   STARTUP IDEAS REPORT  —  {now}  —  {len(posts)} ideas\n")
        f.write("=" * (sum(col_w) + len(col_w) * 3 + 1) + "\n\n")

        f.write(sep + "\n")
        f.write(row_line(["#", "Startup Name", "Description", "Upvotes"], col_w) + "\n")
        f.write(sep + "\n")

        for i, (title, description, upvotes) in enumerate(posts, start=1):
            f.write(row_line([i, title, description, upvotes], col_w) + "\n")
            f.write(sep + "\n")

        total_upvotes = sum(u for _, _, u in posts if isinstance(u, int))
        f.write(f"\n  Total Ideas: {len(posts)}   |   Total Upvotes: {total_upvotes:,}\n")


